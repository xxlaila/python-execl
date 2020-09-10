# -*- coding: utf-8 -*-
"""
@File    : auto_execl.py
@Time    : 2020/9/4 8:53 上午
@Author  : xxlaila
@Email   : admin@xxlaila.cn
@Software: PyCharm
"""

from openpyxl.styles import PatternFill, Font, Alignment, Side, Border, NamedStyle
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from os import remove, path
import datetime
import os
import pandas as pd


class XlsxSaver:

    def __init__(self, sheet_name, datas, fields):
        self.datas = datas
        self.fields = fields
        self.file_path = '/opt/'
        self.pathfile = self.file_path + self.getYesterday() + '.xlsx'
        if os.path.isfile(self.pathfile):
            self.Wb = load_workbook(self.pathfile)
            self.Ws = self.Wb.create_sheet(title=sheet_name, index=-1)
        else:
            self.Wb = openpyxl.Workbook()
            self.Ws = self.Wb.active
            self.Ws.title = sheet_name

        self.WriteExecl()

    def getYesterday(self):
        yesterday = (datetime.date.today() + datetime.timedelta(days=-1)).strftime("%Y.%m.%d")
        return yesterday

    def Merge_Cell(self):
        """
        合并第一行
        :return:
        """
        for row in self.Wb.sheetnames:
            active_sheet = self.Wb[row]
            maxs_rows = active_sheet.max_row
            maxs_cols = active_sheet.max_column

            for col in list(active_sheet.rows)[0]:
                col.merge_cells(maxs_cols)

                # col.merge_cells(start_row=2, start_column=6, end_row=3, end_column=8)
                # col.cell(2, 6).value = '合并三个单元格'

    def UnMerge_Cell(self):
        """
        取消合并
        :return:
        """
        for row in self.Wb.sheetnames:
            active_sheet = self.Wb[row]
            maxs_rows = active_sheet.max_row
            maxs_cols = active_sheet.max_column

            for col in list(active_sheet.rows)[0]:
                col.unmerge_cells(maxs_cols)

    def ReadFrom(self):
        """
        设置表头信息
        :return:
        """
        yellow_fill = PatternFill(fill_type="solid", start_color="87CEFA", end_color="87CEFA")
        """
        fill_type=None,    # 设置填充样式:('darkGrid', 'darkTrellis', 'darkHorizontal', 'darkGray', 'lightDown', 'lightGray', 
        'solid', 'lightGrid', 'gray125', 'lightHorizontal', 'lightTrellis', 'darkDown', 'mediumGray', 'gray0625', 'darkUp', 
        'darkVertical', 'lightVertical', 'lightUp')
        start_color=None   # 设置填充颜色
        end_color=None     # 背景色
        """
        fant1 = Font(name='宋体', size=13, italic=False, color="FF000000", bold=True)
        font2 = Font(name='微软雅黑', size=11)
        """
        name:字体名称
        size: 字体大小
        color:颜色通常是RGB或aRGB十六进制值
        bold:加粗（bool）True/False
        italic:倾斜(bool) True/False
        shadow：阴影（bool）
        underline：下划线（‘doubleAccounting’, ‘single’, ‘double’, ‘singleAccounting’）
        charset:字符集(int)
        strike:删除线(bool) True/False
        """
        alignment = Alignment(horizontal='center', vertical='center')
        """
        # 剧中
        horizontal:水平对齐('centerContinuous', 'general', 'distributed',
                            'left', 'fill', 'center', 'justify', 'right'
                            水平向方: center 靠左：left 靠右: right)
        vertical:垂直对齐（'distributed', 'top', 'center', 'justify', 'bottom'）
        """
        # 设置边框样式
        side = Side(style="thin", color="FFFFFF")
        frames = Border(left=side, right=side, top=side, bottom=side)
        """
        置边对象(四个边可以是一样的也可以不同,如果不同就创建多个Side对象), 设置边框对象(left、right、top、bottom表示的是边框的四个边，四个边使用的是一个边对象)
        style:边框线的风格{'dotted','slantDashDot','dashDot','hair','mediumDashDot',
        'dashed','mediumDashed','thick','dashDotDot','medium',
        'double','thin','mediumDashDotDot'}
        top（上）,bottom（下）,left（左）,right（右）:必须是 Side类型
        diagonal: 斜线 side类型 
        diagonalDownd: 右斜线 bool
        diagonalDown: 左斜线 bool
        """
        # 边框
        line_t = Side(style='thin', color='000000')  # 细边框
        line_m = Side(style='medium', color='000000')  # 粗边框
        # 与标题相邻的边设置与标题一样
        border1 = Border(top=line_m, bottom=line_t, left=line_t, right=line_t)
        border2 = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)

        # 将样式打包命名
        sty1 = NamedStyle(name='sty1', font=fant1, fill=yellow_fill, border=border1, alignment=alignment)
        sty2 = NamedStyle(name='sty2', font=font2, border=border2, alignment=alignment)


        for row in self.Wb.sheetnames:
            active_sheet = self.Wb[row]
            maxs_rows = active_sheet.max_row
            maxs_cols = active_sheet.max_column
            for col in list(active_sheet.rows)[0]:
                col.alignment = alignment
                col.fill = yellow_fill
                col.font = fant1

            active_sheet.row_dimensions[1].height = 28      # 行高

            # 边框
            """
            for r in range(1,maxs_rows+1):
                for c in range(1,maxs_cols):
                    if r == 2:
                        active_sheet.cell(r,c).style = sty1
                    else:
                        active_sheet.cell(r,c).style = sty2
            """

        # self.Wb.save(self.pathfile)

    def WriteExecl(self):
        """
        写数据，1、for 第一行， 2、for 数据
        :return:
        """
        for col in range(len(self.fields)):
            _ = self.Ws.cell(row=1, column=col + 1, value=u'%s' % self.fields[col][0:])

        for row in range(len(self.datas)):
            for col in range(len(self.fields)):
                _ = self.Ws.cell(row=row + 2, column=col + 1, value=u'%s' % self.datas[row][col])

    def auto_width(self):
        """
        自动列宽
        :return:
        """
        # 每个表
        for row in self.Wb.sheetnames:
            active_sheet = self.Wb[row]
            col_width = []
            maxs_rows = active_sheet.max_row
            maxs_cols = active_sheet.max_column
            i = 0
            # 每列
            for col in active_sheet.columns:
                # 每行
                for j in range(len(col)):
                    if j == 0:
                        # 数组增加一个元素
                        col_width.append(len(str(col[j].value)))
                    else:
                        # 获得每列中的内容的最大宽度
                        # print("不为0: %s" % col_width)
                        if max(col_width) < len(str(col[j].value)):
                            col_width[i] = len(str(col[j].value))
                            
                i = i + 1
            # 设置列宽
            for i in range(len(col_width)):
                # 根据列的数字返回字母，提取最大的列宽
                col_letter = get_column_letter(i+1)
                if col_width[i] > 100:
                    active_sheet.column_dimensions[col_letter].width = 100

                elif col_width[i] > 75 and col_width[i] < 90:
                    active_sheet.column_dimensions[col_letter].width = 88
                elif col_width[i] > 55 and col_width[i] < 75:
                    active_sheet.column_dimensions[col_letter].width = 68
                elif max(col_width) > 35 and col_width[i] < 55:
                    active_sheet.column_dimensions[col_letter].width = 48
                elif col_width[i] > 15 and col_width[i] < 35:
                    active_sheet.column_dimensions[col_letter].width = 28
                else:
                # elif col_width[i] > 8 and  col_width[i] < 15:
                #     active_sheet.column_dimensions[col_letter].width = col_width[i] + 2
                    active_sheet.column_dimensions[col_letter].width = 13.8

    def SaveExecl(self):
        """
        保存数据
        :return: 文件名
        """
        self.auto_width()
        self.ReadFrom()
        newworkbook = self.Wb.save(self.pathfile)
        self.Wb.close()
        return newworkbook


# if __name__ == "__main__":
#     sheet_name = "测试"
#     datas = [["a", "b", "c", "a", "r"], ["a", "b", "c", "a", "r"]]
#     fields = ['环境', '项目', '类型', '占用空间(gb)', '文档数量']
#     zxc = XlsxSaver(sheet_name, datas, fields).Merge_Cell()
